import openpyxl
import json
import os
from datetime import datetime, timedelta

# --- Helper: Corporate Rounding Logic ---
def round_to_quarter(hours):
    if hours is None: return 0.0
    return round(hours * 4) / 4

# --- 1. The Configuration Helper ---
def create_default_config():
    default_mapping = {
        "template_path": "timesheet_template.xlsx",
        "output_folder": "completed_timesheets",
        "rounding_enabled": True
    }
    with open("config.json", "w") as f:
        json.dump(default_mapping, f, indent=4)

# --- 2. The Main Export Engine ---
def export_timesheet(data):
    if not os.path.exists("config.json"):
        create_default_config()
        
    with open("config.json", "r") as f:
        config = json.load(f)
    
    template_path = config["template_path"]
    should_round = config.get("rounding_enabled", True)

    # --- SMART TEMPLATE CREATION ---
    # If the template doesn't exist, we create it with THIS WEEK'S dates.
    if not os.path.exists(template_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # 1. Calculate Monday of the current week
        today = datetime.now()
        start_of_week = today - timedelta(days=today.weekday()) # Monday
        
        # 2. Generate Headers (e.g., "Mon 12/11")
        headers = ["Project Name", ""] # Col A and B
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        
        for i, day_name in enumerate(days):
            date_obj = start_of_week + timedelta(days=i)
            date_str = date_obj.strftime("%m/%d")
            headers.append(f"{day_name} {date_str}")

        # 3. Write Headers to Row 4
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=4, column=col_num, value=header)
            
        wb.save(template_path)

    try:
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active 

        # Write Employee Name
        sheet["B2"] = data.get("employee_name", "Mom")

        # --- DYNAMIC ROW WRITING ---
        start_row = 5
        
        for index, project in enumerate(data.get("projects", [])):
            current_row = start_row + index
            
            # Write Project Name (Column A)
            sheet.cell(row=current_row, column=1, value=project["name"])
            
            # Write Hours (Columns C through I)
            days = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
            for day_idx, day_key in enumerate(days):
                hours = project.get(day_key, 0)
                if should_round:
                    hours = round_to_quarter(hours)
                
                # Column C is index 3
                sheet.cell(row=current_row, column=3 + day_idx, value=hours if hours > 0 else "")

        # --- TIMESTAMP LOGIC ---
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_filename = f"Timesheet_{timestamp}.xlsx"
        
        output_dir = config.get("output_folder", ".")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        save_path = os.path.join(output_dir, output_filename)
        wb.save(save_path)
        
        return f"Success! Saved to {save_path}"

    except PermissionError:
        return "Error: Please close the Excel file before exporting!"
    except Exception as e:
        return f"An error occurred: {str(e)}"
